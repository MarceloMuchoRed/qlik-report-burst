"""
Weekly per-employee Qlik dashboard burst: for each row in a recipient
spreadsheet, render the Qlik Sense chart filtered to that person and email the
image via the local Outlook desktop app.

Design notes:
  * The per-employee filter is applied through the Qlik Single Integration API
    URL (&select=Field,Value) rather than by driving the filter pane.
  * The Qlik virtual proxy uses Windows authentication
    (/internal_windows_authentication/) — an NTLM/Negotiate challenge (the
    browser credential popup), not a web form. The bot answers that challenge
    with a specific service account via Playwright's http_credentials, so it
    authenticates as that licensed account instead of silently single-signing-on
    as the VM's own (unlicensed) Windows login. Credentials come from
    QLIK_USERNAME/QLIK_PASSWORD or a gitignored qlik_credentials.txt, never from
    this file.
  * Email is sent through the desktop Outlook via COM (no SMTP/OAuth setup).

See README.md for setup and configuration.
"""

import csv
import os
import sys
import time
import unicodedata
from datetime import date
from urllib.parse import quote

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ==========================================================================
# CONFIG
# ==========================================================================

# Qlik host, no trailing slash (e.g. http://10.0.2.5). Include a virtual-proxy
# prefix here if your hub is behind one (e.g. http://10.0.2.5/sales).
TENANT_URL = "http://10.0.2.5"

# App (document) GUID, from the app URL between /app/ and /sheet/.
APP_ID = "48ab9fa2-5dc9-48f1-8b0e-25041e6313bd"

# Capture a single object (OBJECT_ID); if empty, capture the whole sheet
# (SHEET_ID). See README for how to find the ids.
OBJECT_ID = ""
SHEET_ID = "6527c8b7-f73a-4a7c-962c-6f347d52a009"

# Sheet for the email's "detail of your performance" link. Can differ from the
# screenshot's SHEET_ID (which the capture and the "dashboard" link both use).
DETAIL_SHEET_ID = "1266bc38-8212-4401-aa26-b3652bb6483d"

# The filter field, exactly as Qlik names it (case-sensitive).
FILTER_FIELD = "SALESPERSON_ORDER"

# Recipient list (.xlsx or .csv) and its column headers.
RECIPIENTS_FILE = os.path.join(SCRIPT_DIR, "recipients.xlsx")
NAME_COLUMN = "Employee"
EMAIL_COLUMN = "Email"

# Email content placeholders: {name} = employee full name, {date} = run date
# (M/D/YYYY), {dashboard_url} = link to the dashboard sheet (SHEET_ID),
# {detail_url} = link to the detail sheet (DETAIL_SHEET_ID); both are personalized
# to the employee. The image embeds inline via the cid referenced in the body.
EMAIL_SUBJECT = "Sales Report {date}"
# On-screen display width of the embedded image, in pixels. The screenshot is
# captured at high resolution (VIEWPORT_WIDTH x DEVICE_SCALE), so Outlook would
# otherwise show it at its full pixel width. Setting a width downscales it in the
# client for a sensible size while keeping the extra pixels for a crisp render.
# ~850px fits most Outlook reading panes without horizontal scroll.
# NOTE: use only the plain width attribute here. Outlook's desktop (Word) engine
# mishandles CSS like max-width/height:auto on large images and can CROP them;
# the bare width attribute makes it scale height proportionally, no clipping.
EMAIL_IMAGE_WIDTH = 850
EMAIL_HTML_BODY = """
<p>Hi {name},</p>
<p>Here is your dashboard for this week: <a href="{dashboard_url}">Dashboard Sales KPI Performance v.4</a></p>
<p>Here is the detail of your performance: <a href="{detail_url}">View your performance detail</a></p>
<p>Note: these links open in your default browser, which must be signed in to Qlik for them to work.</p>
<p>Any doubts or issues feel free to contact me via email: <a href="mailto:gerson@pennrosefarms.com">gerson@pennrosefarms.com</a></p>
<p>Problems with your login contact our IT team via email: <a href="mailto:it@pennrosefarms.com">it@pennrosefarms.com</a></p>
<p><img src="cid:dashboard_image" width="{img_width}"></p>
<p>Regards</p>
"""

# --- Test / safety switches -------------------------------------------------
REVIEW_MODE = True                                # True = open drafts, don't send
TEST_REDIRECT_EMAIL = "gerson@pennrosefarms.com"  # send all mail here; "" = real recipients
MAX_EMPLOYEES = None                              # cap rows processed; None = all
RUN_ID_PREFLIGHT = False                          # verify app/sheet IDs first; opens an EXTRA engine session
# When True, look up the real list of FILTER_FIELD values in Qlik up front and
# skip any recipient whose name isn't among them. This matters because the Single
# Integration API silently ignores an unknown selection value and renders the
# whole-company dashboard instead of erroring — so an unknown name would be
# emailed company-wide totals. Skipped names are reported at the end. If the
# lookup can't run (e.g. engine unreachable), validation is disabled with a
# warning and everyone is processed as before.
VALIDATE_NAMES = True
# Show a Windows pop-up with the run summary at the end (counts of sent/drafted,
# names skipped as not-found, and any failures). Set False for unattended /
# scheduled runs with no logged-in desktop, where a modal box has nobody to
# click it (the same summary is always printed to the console regardless).
SHOW_SUMMARY_POPUP = True

# Where screenshots are written (next to this script by default).
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "screenshots")

# --- Browser ----------------------------------------------------------------
# "chrome" uses installed Chrome, falling back to Playwright's bundled Chromium.
# Runs in a throwaway profile, so it never touches your real Chrome profile.
BROWSER_CHANNEL = "chrome"
HEADLESS = False              # True for silent scheduled runs
# Screenshot canvas. A larger viewport makes Qlik render its full desktop layout
# (a small viewport triggers a cramped/rescaled "small screen" layout); bump it
# if your dashboard is bigger. DEVICE_SCALE = 2 doubles pixel density for crisp
# images (a 1920x1080 viewport then yields a 3840x2160 shot).
VIEWPORT_WIDTH = 1920
VIEWPORT_HEIGHT = 1080
DEVICE_SCALE = 2
RENDER_SETTLE_SECONDS = 5     # cushion after charts paint (on top of the render wait)
READY_SELECTOR = ""           # CSS selector that appears once drawn; "" = generic

# --- Qlik login (Windows authentication) ------------------------------------
# The Qlik virtual proxy challenges with Windows auth (NTLM/Negotiate). The bot
# answers that challenge with the credentials below (a LICENSED service account,
# e.g. gmrqlik). Do NOT rely on the VM's own Windows login: that signs in
# silently as the VM account, which has no Qlik access pass. If the account is
# domain-joined and plain "gmrqlik" is rejected, try DOMAIN\gmrqlik or
# gmrqlik@domain.
# Credentials are read from env vars QLIK_USERNAME/QLIK_PASSWORD, else a
# gitignored qlik_credentials.txt (KEY=VALUE). They are never stored in this file.
CREDENTIALS_FILE = os.path.join(SCRIPT_DIR, "qlik_credentials.txt")

# ==========================================================================


def load_recipients(path):
    """Return a list of {name, email} dicts from an .xlsx or .csv file."""
    if not os.path.exists(path):
        sys.exit(f"Recipients file not found: {path}")

    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                rows.append(r)
    elif ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({header[i]: row[i] for i in range(len(header))})
    else:
        sys.exit(f"Unsupported recipients file type: {ext} (use .csv or .xlsx)")

    people = []
    for r in rows:
        name = str(r.get(NAME_COLUMN) or "").strip()
        email = str(r.get(EMAIL_COLUMN) or "").strip()
        if name and email:
            people.append({"name": name, "email": email})
    if not people:
        sys.exit(
            f"No rows found. Check that '{NAME_COLUMN}' and '{EMAIL_COLUMN}' "
            f"match the column headers in {path}."
        )
    return people


def read_credentials():
    """Return (username, password) from env vars, then qlik_credentials.txt.
    Returns ("", "") if neither is configured."""
    user = os.environ.get("QLIK_USERNAME", "").strip()
    pwd = os.environ.get("QLIK_PASSWORD", "")
    if user and pwd:
        return user, pwd
    if os.path.exists(CREDENTIALS_FILE):
        try:
            with open(CREDENTIALS_FILE, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, val = line.split("=", 1)
                    key = key.strip().upper()
                    val = val.strip()
                    if key in ("QLIK_USERNAME", "USERNAME") and not user:
                        user = val
                    elif key in ("QLIK_PASSWORD", "PASSWORD") and not pwd:
                        pwd = val
        except OSError:
            pass
    return user, pwd


def build_single_url(employee_name):
    """Build a Single Integration API URL with the employee selection applied."""
    base = TENANT_URL.rstrip("/") + "/single/"
    params = [f"appid={quote(APP_ID, safe='')}"]
    if OBJECT_ID:
        params.append(f"obj={quote(OBJECT_ID, safe='')}")
    elif SHEET_ID:
        params.append(f"sheet={quote(SHEET_ID, safe='')}")
    else:
        sys.exit("Set either OBJECT_ID or SHEET_ID in CONFIG.")
    sel = f"{quote(FILTER_FIELD, safe='')},{quote(str(employee_name), safe='')}"
    params.append(f"select={sel}")
    params.append("opt=nointeraction")  # hide selection bars/toolbars for a clean shot
    return base + "?" + "&".join(params)


def build_detail_url(employee_name, sheet_id):
    """Build a Qlik Sense client deep link that opens the app on `sheet_id` with
    the employee selected in FILTER_FIELD, so the recipient lands on only their
    own results. Mirrors the hub URL shape (…/state/analysis/…/select/FIELD/VALUE)."""
    base = TENANT_URL.rstrip("/")
    value = quote(str(employee_name), safe="")
    return (
        f"{base}/sense/app/{APP_ID}/sheet/{sheet_id}"
        f"/state/analysis/options/clearselections"
        f"/select/{FILTER_FIELD}/{value}"
    )


def _normalize_name(s):
    """Loose key for matching a recipient name to a Qlik field value: case-,
    whitespace-, accent- and trailing-period-insensitive, so "Jose Nunez Jr"
    matches Qlik's "José Núñez Jr.". Applied to BOTH sides, so it doesn't matter
    which one carries the accents. Used only to look up the exact Qlik value; that
    exact value is what gets selected/linked downstream."""
    s = " ".join(str(s).split()).casefold().rstrip(". ")
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def ensure_logged_in(page):
    """Open the hub. The Windows-auth challenge is answered at the HTTP layer by
    the context's http_credentials, so this just navigates and reports where we
    landed. Staying on the /internal_windows_authentication/ endpoint means the
    service-account credentials were rejected."""
    target = TENANT_URL.rstrip("/") + "/hub/"
    print(f"Opening Qlik: {target}")
    try:
        page.goto(target, wait_until="domcontentloaded", timeout=60000)
        time.sleep(2)
    except Exception as e:
        print(f"   ! couldn't load the hub ({e}).")
    print(f"   landed at: {page.url}")
    if "authentication" in (page.url or "").lower():
        print("   ! still on the Windows-auth endpoint — the service-account "
              "credentials look rejected. Check QLIK_USERNAME/QLIK_PASSWORD "
              "(a domain account may need DOMAIN\\user or user@domain form).")


def fetch_valid_field_values(page):
    """Ask the Qlik engine for the full list of values in FILTER_FIELD, so we can
    tell which recipient names actually exist. Returns a set of value strings, or
    None if the lookup couldn't be performed (validation is then skipped).

    Uses the engine JSON-RPC API over a WebSocket opened from inside the already
    authenticated browser page, so the same Windows-auth session cookie is reused
    (no second login). It opens the app, creates a session list object on the
    field, and reads back its values.
    """
    base = TENANT_URL.rstrip("/")
    if base.startswith("https://"):
        ws_url = "wss://" + base[len("https://"):]
    elif base.startswith("http://"):
        ws_url = "ws://" + base[len("http://"):]
    else:
        ws_url = "ws://" + base
    ws_url = ws_url + "/app/" + APP_ID

    js = r"""
    async ({ wsUrl, field }) => {
      const TIMEOUT_MS = 20000;
      const ws = new WebSocket(wsUrl);
      const pending = {};
      let idc = 0;
      ws.onmessage = (ev) => {
        let msg;
        try { msg = JSON.parse(ev.data); } catch (e) { return; }
        if (msg.id && pending[msg.id]) {
          const cb = pending[msg.id];
          delete pending[msg.id];
          cb(msg);
        }
      };
      const opened = new Promise((resolve, reject) => {
        ws.onopen = () => resolve();
        ws.onerror = () => reject(new Error('websocket connection failed'));
      });
      const timeout = new Promise((_, reject) =>
        setTimeout(() => reject(new Error('engine lookup timed out')), TIMEOUT_MS));
      const call = (handle, method, params) => new Promise((resolve, reject) => {
        const id = ++idc;
        pending[id] = (msg) => {
          if (msg.error) reject(new Error((msg.error && msg.error.message) || 'engine error'));
          else resolve(msg.result);
        };
        ws.send(JSON.stringify({ jsonrpc: '2.0', id, handle, method, params }));
      });
      const run = (async () => {
        await opened;
        // A /app/<id> socket opens that app automatically; grab its handle.
        let docHandle = -1;
        try {
          const active = await call(-1, 'GetActiveDoc', []);
          if (active && active.qReturn && typeof active.qReturn.qHandle === 'number'
              && active.qReturn.qHandle >= 0) {
            docHandle = active.qReturn.qHandle;
          }
        } catch (e) { /* fall through to OpenDoc */ }
        if (docHandle < 0) {
          const od = await call(-1, 'OpenDoc', ['APPID_PLACEHOLDER', '', '', '', false]);
          docHandle = od.qReturn.qHandle;
        }
        const obj = await call(docHandle, 'CreateSessionObject', [{
          qInfo: { qType: 'ListObject' },
          qListObjectDef: {
            qDef: { qFieldDefs: [field] },
            qInitialDataFetch: [{ qTop: 0, qLeft: 0, qHeight: 10000, qWidth: 1 }]
          }
        }]);
        const objHandle = obj.qReturn.qHandle;
        const layout = await call(objHandle, 'GetLayout', []);
        const lo = layout.qLayout.qListObject;
        const total = (lo.qSize && typeof lo.qSize.qcy === 'number') ? lo.qSize.qcy : null;
        const values = [];
        for (const dp of (lo.qDataPages || [])) {
          for (const row of dp.qMatrix) {
            for (const cell of row) {
              if (cell && cell.qText !== undefined && !cell.qIsNull) values.push(cell.qText);
            }
          }
        }
        ws.close();
        return { values, total };
      })();
      return await Promise.race([run, timeout]);
    }
    """.replace("APPID_PLACEHOLDER", APP_ID)

    try:
        result = page.evaluate(js, {"wsUrl": ws_url, "field": FILTER_FIELD})
    except Exception as e:
        print(f"   ! couldn't read '{FILTER_FIELD}' values from the Qlik engine "
              f"({e}); name validation is DISABLED for this run.")
        return None

    values = (result or {}).get("values") or []
    total = (result or {}).get("total")
    if not values:
        print(f"   ! the Qlik engine returned no values for '{FILTER_FIELD}'; "
              f"name validation is DISABLED for this run.")
        return None
    if total and len(values) >= 10000 and total > len(values):
        print(f"   ! '{FILTER_FIELD}' has {total} values but only the first "
              f"{len(values)} were read; some names may be flagged wrongly.")
    return set(values)


def verify_ids(page):
    """Best-effort preflight: confirm APP_ID opens and the sheet IDs exist, over
    the same authenticated engine websocket used for the name lookup. Returns None
    if the check itself couldn't run (caller then warns and continues rather than
    blocking a good run); otherwise {"appOk": bool, "sheets": [ids] or None}."""
    base = TENANT_URL.rstrip("/")
    if base.startswith("https://"):
        ws_url = "wss://" + base[len("https://"):]
    elif base.startswith("http://"):
        ws_url = "ws://" + base[len("http://"):]
    else:
        ws_url = "ws://" + base
    ws_url = ws_url + "/app/" + APP_ID

    js = r"""
    async ({ wsUrl, appId }) => {
      const TIMEOUT_MS = 20000;
      const ws = new WebSocket(wsUrl);
      const pending = {};
      let idc = 0;
      ws.onmessage = (ev) => {
        let m; try { m = JSON.parse(ev.data); } catch (e) { return; }
        if (m.id && pending[m.id]) { const cb = pending[m.id]; delete pending[m.id]; cb(m); }
      };
      const opened = new Promise((res, rej) => {
        ws.onopen = () => res();
        ws.onerror = () => rej(new Error('websocket connection failed'));
      });
      const timeout = new Promise((_, rej) =>
        setTimeout(() => rej(new Error('engine check timed out')), TIMEOUT_MS));
      const call = (handle, method, params) => new Promise((res, rej) => {
        const id = ++idc;
        pending[id] = (m) => {
          if (m.error) rej(new Error((m.error && m.error.message) || 'engine error'));
          else res(m.result);
        };
        ws.send(JSON.stringify({ jsonrpc: '2.0', id, handle, method, params }));
      });
      const run = (async () => {
        await opened;
        let docHandle = -1;
        try {
          const a = await call(-1, 'GetActiveDoc', []);
          if (a && a.qReturn && typeof a.qReturn.qHandle === 'number' && a.qReturn.qHandle >= 0)
            docHandle = a.qReturn.qHandle;
        } catch (e) { /* fall through to OpenDoc */ }
        if (docHandle < 0) {
          try {
            const od = await call(-1, 'OpenDoc', [appId, '', '', '', false]);
            docHandle = od.qReturn.qHandle;
          } catch (e) { return { appOk: false, sheets: null }; }
        }
        try {
          const r = await call(docHandle, 'GetObjects',
            [{ qOptions: { qTypes: ['sheet'], qIncludeSessionObjects: false, qData: {} } }]);
          const list = (r && r.qList) || [];
          const sheets = list.map(e => e.qInfo && e.qInfo.qId).filter(Boolean);
          return { appOk: true, sheets };
        } catch (e) {
          return { appOk: true, sheets: null };
        }
      })();
      try { return await Promise.race([run, timeout]); }
      finally { try { ws.close(); } catch (e) {} }
    }
    """
    try:
        return page.evaluate(js, {"wsUrl": ws_url, "appId": APP_ID})
    except Exception as e:
        print(f"   ! couldn't verify app/sheet IDs ({e}); continuing without the check.")
        return None


def capture(page, employee_name, out_path):
    """Load the filtered single view, wait for the Qlik objects to finish
    rendering, then screenshot it."""
    url = build_single_url(employee_name)
    # domcontentloaded (not networkidle): Qlik holds a websocket open so the page
    # never goes fully idle; readiness is enforced below by the selector wait and
    # the object-count-stable loop, so waiting on networkidle only adds dead time.
    page.goto(url, wait_until="domcontentloaded")
    selector = READY_SELECTOR or ".qv-object"
    try:
        page.wait_for_selector(selector, timeout=30000)
    except Exception:
        print(f"   ! '{selector}' never appeared for {employee_name}; "
              f"screenshotting whatever is on screen.")

    # Qlik lazy-loads in two stages: the object *containers* (`.qv-object`) appear
    # first, then each chart paints into a <canvas>/<svg> a beat later. Waiting on
    # the container count alone can fire while a chart is still blank — the KPI
    # text is present but the main chart hasn't drawn (the empty-chart bug). So
    # wait until BOTH the container count and the painted chart-element count stop
    # growing, for two consecutive reads, before the fixed settle.
    prev = (-1, -1)
    stable = 0
    for _ in range(40):  # up to ~20s
        objs = page.locator(selector).count()
        arts = page.locator(f"{selector} canvas, {selector} svg").count()
        if objs > 0 and (objs, arts) == prev:
            stable += 1
            if stable >= 2:
                break
        else:
            stable = 0
        prev = (objs, arts)
        page.wait_for_timeout(500)
    time.sleep(RENDER_SETTLE_SECONDS)

    element = None
    if OBJECT_ID:
        element = page.query_selector(".qv-object") or page.query_selector(selector)
    if element:
        element.screenshot(path=out_path)
    else:
        page.screenshot(path=out_path, full_page=True)


def send_email(name, to_address, image_path):
    """Create (and optionally send) an email through the desktop Outlook app."""
    import win32com.client  # from pywin32
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)  # 0 = olMailItem
    mail.To = to_address
    today = date.today()
    mail.Subject = EMAIL_SUBJECT.format(date=f"{today.month}/{today.day}/{today.year}")

    attachment = mail.Attachments.Add(os.path.abspath(image_path))
    try:
        # PR_ATTACH_CONTENT_ID: lets HTMLBody reference the image with cid:
        attachment.PropertyAccessor.SetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x3712001F",
            "dashboard_image",
        )
    except Exception:
        pass  # if the cid fails, the image still rides along as an attachment
    mail.HTMLBody = EMAIL_HTML_BODY.format(
        name=name,
        dashboard_url=build_detail_url(name, SHEET_ID),
        detail_url=build_detail_url(name, DETAIL_SHEET_ID),
        img_width=EMAIL_IMAGE_WIDTH,
    )

    if REVIEW_MODE:
        mail.Display(False)  # open as a draft; does not send
        return "drafted"
    mail.Send()
    return "sent"


def show_summary(succeeded, skipped, failed):
    """Print, and (if enabled) pop up, a summary of the run. `succeeded` is a list
    of (name, status) where status is 'sent' or 'drafted'; `skipped` is the list
    of not-found person dicts; `failed` is a list of (name, error)."""
    sent = [n for n, s in succeeded if s == "sent"]
    drafted = [n for n, s in succeeded if s == "drafted"]

    lines = []
    if sent:
        lines.append(f"Sent: {len(sent)}")
    if drafted:
        lines.append(f"Drafted (not sent, REVIEW_MODE): {len(drafted)}")
    if skipped:
        lines.append(f"Skipped - name not found in Qlik: {len(skipped)}")
        for p in skipped:
            lines.append(f"   - {p['name']}")
    if failed:
        lines.append(f"Failed: {len(failed)}")
        for name, err in failed:
            lines.append(f"   - {name}: {err}")
    if not lines:
        lines.append("No recipients were processed.")
    text = "\n".join(lines)

    print("\n=== Summary ===")
    print(text)

    if SHOW_SUMMARY_POPUP:
        try:
            import ctypes
            # MB_OK | MB_ICONINFORMATION | MB_SETFOREGROUND | MB_TOPMOST
            flags = 0x40 | 0x10000 | 0x40000
            ctypes.windll.user32.MessageBoxW(
                0, text, "Qlik report burst - summary", flags)
        except Exception as e:
            print(f"   ! couldn't show the summary pop-up ({e}).")


def launch_browser(p):
    """Launch the configured browser in a throwaway profile, falling back to
    Playwright's bundled Chromium if the installed Chrome channel isn't found."""
    try:
        return p.chromium.launch(channel=BROWSER_CHANNEL, headless=HEADLESS)
    except Exception as e:
        print(f"   ! couldn't launch '{BROWSER_CHANNEL}' ({e}); "
              f"falling back to Playwright's bundled Chromium.")
        return p.chromium.launch(headless=HEADLESS)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    people = load_recipients(RECIPIENTS_FILE)
    if MAX_EMPLOYEES is not None:
        people = people[:MAX_EMPLOYEES]

    username, password = read_credentials()
    if not (username and password):
        sys.exit(
            "No Qlik credentials set. The server uses Windows authentication, so "
            "the bot must send a service account. Set QLIK_USERNAME/QLIK_PASSWORD "
            "(env vars) or fill qlik_credentials.txt."
        )

    print(f"Processing {len(people)} employee(s). "
          f"REVIEW_MODE={REVIEW_MODE}  "
          f"redirect={'ON -> ' + TEST_REDIRECT_EMAIL if TEST_REDIRECT_EMAIL else 'off'}")
    print(f"Qlik login (Windows auth): {username}")

    print("Starting the browser engine...", flush=True)
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        browser = launch_browser(p)
        # Answer the Qlik proxy's Windows-auth (NTLM/Negotiate) challenge with the
        # service account, rather than the VM's ambient login.
        context = browser.new_context(
            http_credentials={"username": username, "password": password},
            viewport={"width": VIEWPORT_WIDTH, "height": VIEWPORT_HEIGHT},
            device_scale_factor=DEVICE_SCALE,
        )
        page = context.new_page()

        ensure_logged_in(page)

        # Preflight: verify the app and sheet IDs exist before doing any work.
        # OFF by default — it opens an EXTRA engine session, which on an access-
        # limited service account (e.g. gmrqlik) can starve the first capture's
        # session and surface "a resource could not be found". Flip on to use it.
        if RUN_ID_PREFLIGHT:
            print("Verifying app/sheet IDs...")
            ids = verify_ids(page)
            if ids is not None:
                if ids.get("appOk") is False:
                    sys.exit(f"App ID not found in Qlik: {APP_ID}. Check APP_ID in CONFIG.")
                sheets = ids.get("sheets")
                if sheets is not None:
                    missing = [s for s in {SHEET_ID, DETAIL_SHEET_ID} if s and s not in sheets]
                    if missing:
                        sys.exit("Sheet ID(s) not found in the app: " + ", ".join(missing)
                                 + ". Check SHEET_ID / DETAIL_SHEET_ID in CONFIG.")
                    print(f"   app + {len(sheets)} sheet(s) OK.")
                else:
                    print("   app opened, but couldn't list sheets; skipping sheet check.")

        # Look up the real field values so unknown names are caught before they
        # render (and get emailed) the whole-company dashboard, and so a recipient
        # spelled slightly differently (e.g. "Jr" vs Qlik's "Jr.") resolves to the
        # exact Qlik value. None = lookup unavailable, so we don't enforce it.
        canon_map = None  # normalized name -> exact Qlik field value
        if VALIDATE_NAMES:
            print(f"Checking recipient names against Qlik '{FILTER_FIELD}'...")
            valid_values = fetch_valid_field_values(page)
            if valid_values is not None:
                canon_map = {}
                for v in valid_values:
                    canon_map.setdefault(_normalize_name(v), v)
                print(f"   {len(valid_values)} value(s) found in Qlik.")

        skipped = []
        succeeded = []
        failed = []
        for i, person in enumerate(people, 1):
            name = person["name"]
            recipient = TEST_REDIRECT_EMAIL or person["email"]

            # Resolve to the exact Qlik value (handles case/spacing/trailing-period
            # differences); that exact value is what we select and link, so a
            # near-miss never silently shoots the whole-company dashboard.
            qlik_name = name
            if canon_map is not None:
                resolved = canon_map.get(_normalize_name(name))
                if resolved is None:
                    print(f"[{i}/{len(people)}] {name} -> SKIPPED "
                          f"(not found in Qlik '{FILTER_FIELD}'); no email sent.")
                    skipped.append(person)
                    continue
                if resolved != name:
                    print(f"   note: '{name}' matched Qlik value '{resolved}'.")
                qlik_name = resolved

            safe = "".join(c for c in qlik_name if c.isalnum() or c in " _-").strip()
            img = os.path.join(OUTPUT_DIR, f"{i:02d}_{safe or 'employee'}.png")

            print(f"[{i}/{len(people)}] {qlik_name} -> {recipient}")
            try:
                capture(page, qlik_name, img)
                status = send_email(qlik_name, recipient, img)
                print(f"   screenshot saved: {img}")
                print(f"   email {status}")
                succeeded.append((name, status))
            except Exception as e:
                print(f"   ! FAILED for {name}: {e}")
                failed.append((name, str(e)))

        context.close()
        browser.close()

    show_summary(succeeded, skipped, failed)
    print("\nDone. Review the drafts in Outlook (REVIEW_MODE) or check Sent.")


if __name__ == "__main__":
    main()
